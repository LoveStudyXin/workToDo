"""
生成Excel格式的测试报告
运行方式: cd backend && python tests/generate_report.py
"""
import subprocess
import sys
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# 测试用例定义：模块 -> [(方法名, 用例名称, 测试描述)]
TEST_CASES = {
    "用户认证": [
        ("test_register_success", "用户注册", "验证新用户能够成功注册账号"),
        ("test_register_duplicate_username", "重复用户名注册", "验证使用已存在的用户名注册时返回错误"),
        ("test_register_duplicate_email", "重复邮箱注册", "验证使用已存在的邮箱注册时返回错误"),
        ("test_login_success", "用户登录", "验证使用正确的用户名密码能够成功登录"),
        ("test_login_wrong_username", "用户名不存在", "验证用户名不存在时提示\"该用户名未注册，请先注册\""),
        ("test_login_wrong_password", "密码错误", "验证密码错误时提示\"密码错误，请检查\""),
        ("test_get_current_user", "获取当前用户", "验证登录后能够获取当前用户信息"),
    ],
    "任务管理": [
        ("test_create_todo", "创建任务", "验证能够成功创建新任务"),
        ("test_get_todos", "获取任务列表", "验证能够获取用户的任务列表"),
        ("test_get_todos_by_status", "按状态筛选", "验证能够按状态筛选任务"),
        ("test_get_today_todos", "今日任务", "验证能够获取今天到期和进行中的任务"),
        ("test_update_todo", "更新任务", "验证能够更新任务的标题、状态等信息"),
        ("test_delete_todo", "删除任务", "验证能够删除任务"),
        ("test_complete_todo", "完成任务", "验证完成任务时状态自动变为completed，进度变为100%"),
    ],
    "工作日志": [
        ("test_create_work_log", "创建日志", "验证能够为任务创建工作日志"),
        ("test_get_work_logs", "获取日志列表", "验证能够获取工作日志列表"),
        ("test_get_work_logs_by_date", "按日期筛选", "验证能够按日期范围筛选工作日志"),
        ("test_update_work_log", "更新日志", "验证能够更新工作日志内容"),
        ("test_delete_work_log", "删除日志", "验证能够删除工作日志"),
    ],
    "报告生成": [
        ("test_preview_report", "预览报告", "验证能够预览日报/周报/月报内容"),
        ("test_generate_report", "生成报告", "验证能够生成并保存报告"),
        ("test_get_reports", "获取报告列表", "验证能够获取已生成的报告列表"),
        ("test_get_presentation_styles", "演示文稿风格", "验证能够获取可用的PPT风格列表"),
    ],
    "模板管理": [
        ("test_get_templates", "获取模板列表", "验证能够获取系统模板和用户自定义模板"),
        ("test_copy_template", "复制模板", "验证能够将系统模板复制为自己的模板"),
        ("test_update_own_template", "修改模板", "验证能够修改自己的模板"),
        ("test_cannot_update_system_template", "系统模板保护", "验证无法修改系统默认模板"),
        ("test_delete_own_template", "删除模板", "验证能够删除自己的模板"),
    ],
    "权限控制": [
        ("test_unauthorized_access", "未授权访问", "验证未登录时访问接口返回401错误"),
        ("test_access_other_user_todo", "数据隔离", "验证用户无法访问其他用户的任务数据"),
    ],
}


def run_tests():
    """运行pytest并获取结果"""
    result = subprocess.run(
        [sys.executable, "-m", "pytest", "tests/test_api.py", "-v", "--tb=no"],
        capture_output=True,
        text=True,
        cwd="/Users/xieyuxin/Downloads/工作todo/backend"
    )
    return result.stdout + result.stderr


def parse_test_results(output):
    """解析pytest输出，返回 {方法名: 'PASS'/'FAIL'} """
    results = {}
    for line in output.split('\n'):
        # 匹配格式: tests/test_api.py::TestAuth::test_register_success PASSED
        match = re.search(r'::(\w+)::(\w+)\s+(PASSED|FAILED)', line)
        if match:
            method_name = match.group(2)
            status = 'PASS' if match.group(3) == 'PASSED' else 'FAIL'
            results[method_name] = status
    return results


def create_excel_report(test_results, output_path, version="v1.0.0"):
    """创建Excel测试报告"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试报告"

    # 样式定义
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    module_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ===== 标题区域 =====
    ws.merge_cells('A1:F1')
    ws['A1'] = "工作TODO应用 - API测试报告"
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws['A3'] = "测试日期:"
    ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A4'] = "测试版本:"
    ws['B4'] = version
    ws['A5'] = "测试环境:"
    ws['B5'] = "Python 3.13 + FastAPI + pytest"

    # ===== 统计区域 =====
    total = sum(len(cases) for cases in TEST_CASES.values())
    passed = sum(1 for r in test_results.values() if r == 'PASS')
    failed = total - passed

    ws['D3'] = "通过:"
    ws['E3'] = passed
    ws['E3'].fill = pass_fill
    ws['F3'] = f"({passed/total*100:.0f}%)"

    ws['D4'] = "失败:"
    ws['E4'] = failed
    if failed > 0:
        ws['E4'].fill = fail_fill
    ws['F4'] = f"({failed/total*100:.0f}%)"

    ws['D5'] = "总计:"
    ws['E5'] = total

    # ===== 表头 =====
    headers = ['序号', '模块', '测试用例', '测试描述', '预期结果', '测试结果']
    row = 7
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    # ===== 测试用例详情 =====
    row = 8
    idx = 1
    for module_name, cases in TEST_CASES.items():
        for method_name, case_name, description in cases:
            status = test_results.get(method_name, '未执行')

            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = center

            ws.cell(row=row, column=2, value=module_name).border = border
            ws.cell(row=row, column=2).alignment = center
            ws.cell(row=row, column=2).fill = module_fill

            ws.cell(row=row, column=3, value=case_name).border = border
            ws.cell(row=row, column=3).alignment = center

            ws.cell(row=row, column=4, value=description).border = border
            ws.cell(row=row, column=4).alignment = left

            ws.cell(row=row, column=5, value="接口返回正确响应").border = border
            ws.cell(row=row, column=5).alignment = center

            result_cell = ws.cell(row=row, column=6, value=status)
            result_cell.border = border
            result_cell.alignment = center
            if status == 'PASS':
                result_cell.fill = pass_fill
                result_cell.font = Font(color="006100")
            elif status == 'FAIL':
                result_cell.fill = fail_fill
                result_cell.font = Font(color="9C0006")

            row += 1
            idx += 1

    # ===== 调整列宽 =====
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12

    # 设置行高
    for r in range(8, row):
        ws.row_dimensions[r].height = 25

    wb.save(output_path)
    return passed, failed, total


def get_version():
    """获取版本号，优先从git tag获取，否则使用默认版本"""
    try:
        result = subprocess.run(
            ["git", "describe", "--tags", "--abbrev=0"],
            capture_output=True,
            text=True,
            cwd="/Users/xieyuxin/Downloads/工作todo/backend"
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()
    except:
        pass
    return "v1.0.0"


def main():
    print("正在运行测试...")
    output = run_tests()

    print("正在解析测试结果...")
    test_results = parse_test_results(output)

    print(f"解析到 {len(test_results)} 个测试结果")

    # 生成带日期和版本的文件名
    test_date = datetime.now().strftime("%Y-%m-%d")
    version = get_version()
    output_path = f"/Users/xieyuxin/Downloads/工作todo/backend/tests/TEST_REPORT_{test_date}_{version}.xlsx"

    passed, failed, total = create_excel_report(test_results, output_path, version)

    print(f"\n{'='*50}")
    print(f"测试报告已生成: {output_path}")
    print(f"{'='*50}")
    print(f"版本: {version}")
    print(f"日期: {test_date}")
    print(f"通过: {passed}")
    print(f"失败: {failed}")
    print(f"总计: {total}")
    print(f"通过率: {passed/total*100:.0f}%")


if __name__ == "__main__":
    main()
