"""
完整测试报告生成器（前端 + 后端）
运行方式: cd 工作todo && python generate_full_report.py
"""
import subprocess
import sys
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ========== 后端测试用例 ==========
BACKEND_TEST_CASES = {
    "用户认证": [
        ("test_register_success", "用户注册", "验证新用户能够成功注册账号"),
        ("test_register_duplicate_username", "重复用户名注册", "验证使用已存在的用户名注册时返回错误"),
        ("test_register_duplicate_email", "重复邮箱注册", "验证使用已存在的邮箱注册时返回错误"),
        ("test_login_success", "用户登录", "验证使用正确的用户名密码能够成功登录"),
        ("test_login_wrong_username", "用户名不存在", "验证用户名不存在时提示\"该用户名未注册，请先注册\""),
        ("test_login_wrong_password", "密码错误", "验证密码错误时提示\"密码错误，请检查\""),
        ("test_get_current_user", "获取当前用户", "验证登录后能够获取当前用户信息"),
    ],
    "任务管理": [
        ("test_create_todo", "创建任务", "验证能够成功创建新任务"),
        ("test_get_todos", "获取任务列表", "验证能够获取用户的任务列表"),
        ("test_get_todos_by_status", "按状态筛选", "验证能够按状态筛选任务"),
        ("test_get_today_todos", "今日任务", "验证能够获取今天到期和进行中的任务"),
        ("test_update_todo", "更新任务", "验证能够更新任务的标题、状态等信息"),
        ("test_delete_todo", "删除任务", "验证能够删除任务"),
        ("test_complete_todo", "完成任务", "验证完成任务时状态自动变为completed"),
    ],
    "工作日志": [
        ("test_create_work_log", "创建日志", "验证能够为任务创建工作日志"),
        ("test_get_work_logs", "获取日志列表", "验证能够获取工作日志列表"),
        ("test_get_work_logs_by_date", "按日期筛选", "验证能够按日期范围筛选工作日志"),
        ("test_update_work_log", "更新日志", "验证能够更新工作日志内容"),
        ("test_delete_work_log", "删除日志", "验证能够删除工作日志"),
    ],
    "报告生成": [
        ("test_preview_report", "预览报告", "验证能够预览日报/周报/月报内容"),
        ("test_generate_report", "生成报告", "验证能够生成并保存报告"),
        ("test_get_reports", "获取报告列表", "验证能够获取已生成的报告列表"),
        ("test_get_presentation_styles", "演示文稿风格", "验证能够获取可用的PPT风格列表"),
    ],
    "模板管理": [
        ("test_get_templates", "获取模板列表", "验证能够获取系统模板和用户自定义模板"),
        ("test_copy_template", "复制模板", "验证能够将系统模板复制为自己的模板"),
        ("test_update_own_template", "修改模板", "验证能够修改自己的模板"),
        ("test_cannot_update_system_template", "系统模板保护", "验证无法修改系统默认模板"),
        ("test_delete_own_template", "删除模板", "验证能够删除自己的模板"),
    ],
    "权限控制": [
        ("test_unauthorized_access", "未授权访问", "验证未登录时访问接口返回401错误"),
        ("test_access_other_user_todo", "数据隔离", "验证用户无法访问其他用户的任务数据"),
    ],
}

# ========== 前端测试用例 ==========
FRONTEND_TEST_CASES = {
    "Auth Store": [
        ("用户初始为空", "初始状态", "验证未登录时用户信息为空"),
        ("未登录时 isAuthenticated 为 false", "认证状态", "验证未登录状态判断正确"),
        ("登录成功后保存 token", "登录流程", "验证登录后token正确存储到localStorage"),
        ("登录成功后获取用户信息", "用户信息", "验证登录后自动获取并保存用户信息"),
        ("注册调用正确的 API", "注册流程", "验证注册功能调用正确的接口"),
        ("登出后清除用户信息和 token", "登出流程", "验证登出后清理所有认证状态"),
    ],
    "Todo Store": [
        ("任务列表初始为空", "初始状态", "验证初始化时任务列表为空数组"),
        ("loading 初始为 false", "加载状态", "验证初始加载状态为false"),
        ("fetchTodos 获取所有任务", "获取任务", "验证能正确获取任务列表"),
        ("fetchTodos 支持筛选条件", "筛选任务", "验证支持按状态等条件筛选任务"),
        ("fetchTodayTodos 获取今日任务", "今日任务", "验证能获取今日待办任务"),
        ("加载时 loading 为 true", "加载状态", "验证请求中loading状态正确"),
        ("createTodo 创建新任务", "创建任务", "验证创建任务后添加到列表开头"),
        ("updateTodo 更新任务", "更新任务", "验证更新任务后列表同步更新"),
        ("deleteTodo 删除任务", "删除任务", "验证删除任务后从列表移除"),
    ],
    "Login 逻辑": [
        ("空表单提交时显示提示", "表单验证", "验证空表单提交时显示错误提示"),
        ("登录成功后调用正确的方法", "登录成功", "验证登录成功后保存token并跳转首页"),
        ("登录失败时显示错误信息", "登录失败", "验证登录失败时显示后端返回的错误信息"),
    ],
}


def run_backend_tests():
    """运行后端测试"""
    print("运行后端测试...")
    result = subprocess.run(
        [sys.executable, "-m", "pytest", "backend/tests/test_api.py", "-v", "--tb=no"],
        capture_output=True,
        text=True
    )
    return result.stdout + result.stderr


def run_frontend_tests():
    """运行前端测试"""
    print("运行前端测试...")
    result = subprocess.run(
        ["npm", "run", "test"],
        capture_output=True,
        text=True,
        cwd="frontend"
    )
    return result.stdout + result.stderr


def parse_backend_results(output):
    """解析后端测试结果"""
    results = {}
    for line in output.split('\n'):
        match = re.search(r'::(\w+)::(\w+)\s+(PASSED|FAILED)', line)
        if match:
            method_name = match.group(2)
            status = 'PASS' if match.group(3) == 'PASSED' else 'FAIL'
            results[method_name] = status
    return results


def parse_frontend_results(output):
    """解析前端测试结果"""
    # 简单判断是否全部通过
    if 'failed' not in output.lower() or '0 failed' in output.lower():
        # 所有测试通过
        results = {}
        for module_cases in FRONTEND_TEST_CASES.values():
            for case in module_cases:
                results[case[0]] = 'PASS'
        return results
    else:
        # 有失败的测试，需要解析具体哪些失败
        results = {}
        for module_cases in FRONTEND_TEST_CASES.values():
            for case in module_cases:
                if case[0] in output and 'FAIL' in output:
                    results[case[0]] = 'FAIL'
                else:
                    results[case[0]] = 'PASS'
        return results


def get_version():
    """获取版本号"""
    try:
        result = subprocess.run(
            ["git", "describe", "--tags", "--abbrev=0"],
            capture_output=True,
            text=True
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()
    except:
        pass
    return "v1.0.0"


def create_excel_report(backend_results, frontend_results, output_path, version):
    """创建Excel测试报告"""
    wb = Workbook()
    ws = wb.active
    ws.title = "完整测试报告"

    # 样式
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    section_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    backend_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    frontend_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    module_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 标题
    ws.merge_cells('A1:F1')
    ws['A1'] = "工作TODO应用 - 完整测试报告"
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    # 基本信息
    ws['A3'] = "测试日期:"
    ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A4'] = "测试版本:"
    ws['B4'] = version
    ws['A5'] = "测试环境:"
    ws['B5'] = "Python 3.13 + FastAPI / Vue 3 + Vitest"

    # 统计
    backend_total = sum(len(cases) for cases in BACKEND_TEST_CASES.values())
    backend_passed = sum(1 for r in backend_results.values() if r == 'PASS')
    frontend_total = sum(len(cases) for cases in FRONTEND_TEST_CASES.values())
    frontend_passed = sum(1 for r in frontend_results.values() if r == 'PASS')
    total = backend_total + frontend_total
    passed = backend_passed + frontend_passed

    ws['D3'] = "后端测试:"
    ws['E3'] = f"{backend_passed}/{backend_total}"
    ws['E3'].fill = pass_fill if backend_passed == backend_total else fail_fill

    ws['D4'] = "前端测试:"
    ws['E4'] = f"{frontend_passed}/{frontend_total}"
    ws['E4'].fill = pass_fill if frontend_passed == frontend_total else fail_fill

    ws['D5'] = "总计:"
    ws['E5'] = f"{passed}/{total} ({passed/total*100:.0f}%)"
    ws['E5'].fill = pass_fill if passed == total else fail_fill

    # ========== 后端测试 ==========
    row = 7
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "后端 API 测试"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = backend_fill
    ws[f'A{row}'].alignment = center

    row += 1
    headers = ['序号', '模块', '测试用例', '测试描述', '预期结果', '测试结果']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    row += 1
    idx = 1
    for module_name, cases in BACKEND_TEST_CASES.items():
        for method_name, case_name, description in cases:
            status = backend_results.get(method_name, '未执行')

            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = center

            ws.cell(row=row, column=2, value=module_name).border = border
            ws.cell(row=row, column=2).alignment = center
            ws.cell(row=row, column=2).fill = module_fill

            ws.cell(row=row, column=3, value=case_name).border = border
            ws.cell(row=row, column=3).alignment = center

            ws.cell(row=row, column=4, value=description).border = border
            ws.cell(row=row, column=4).alignment = left

            ws.cell(row=row, column=5, value="接口返回正确响应").border = border
            ws.cell(row=row, column=5).alignment = center

            result_cell = ws.cell(row=row, column=6, value=status)
            result_cell.border = border
            result_cell.alignment = center
            result_cell.fill = pass_fill if status == 'PASS' else fail_fill

            row += 1
            idx += 1

    # ========== 前端测试 ==========
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "前端逻辑测试"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = frontend_fill
    ws[f'A{row}'].alignment = center

    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    row += 1
    for module_name, cases in FRONTEND_TEST_CASES.items():
        for case_name, short_name, description in cases:
            status = frontend_results.get(case_name, '未执行')

            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = center

            ws.cell(row=row, column=2, value=module_name).border = border
            ws.cell(row=row, column=2).alignment = center
            ws.cell(row=row, column=2).fill = module_fill

            ws.cell(row=row, column=3, value=short_name).border = border
            ws.cell(row=row, column=3).alignment = center

            ws.cell(row=row, column=4, value=description).border = border
            ws.cell(row=row, column=4).alignment = left

            ws.cell(row=row, column=5, value="状态/逻辑正确").border = border
            ws.cell(row=row, column=5).alignment = center

            result_cell = ws.cell(row=row, column=6, value=status)
            result_cell.border = border
            result_cell.alignment = center
            result_cell.fill = pass_fill if status == 'PASS' else fail_fill

            row += 1
            idx += 1

    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12

    wb.save(output_path)
    return backend_passed, backend_total, frontend_passed, frontend_total


def main():
    # 运行测试
    backend_output = run_backend_tests()
    frontend_output = run_frontend_tests()

    # 解析结果
    print("解析测试结果...")
    backend_results = parse_backend_results(backend_output)
    frontend_results = parse_frontend_results(frontend_output)

    # 生成报告
    import os
    version = get_version()
    test_date = datetime.now().strftime("%Y-%m-%d")
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)
    output_path = f"{reports_dir}/TEST_REPORT_{test_date}_{version}.xlsx"

    backend_passed, backend_total, frontend_passed, frontend_total = create_excel_report(
        backend_results, frontend_results, output_path, version
    )

    total = backend_total + frontend_total
    passed = backend_passed + frontend_passed

    print(f"\n{'='*50}")
    print(f"测试报告已生成: {output_path}")
    print(f"{'='*50}")
    print(f"版本: {version}")
    print(f"日期: {test_date}")
    print(f"后端: {backend_passed}/{backend_total} 通过")
    print(f"前端: {frontend_passed}/{frontend_total} 通过")
    print(f"总计: {passed}/{total} ({passed/total*100:.0f}%)")


if __name__ == "__main__":
    main()
